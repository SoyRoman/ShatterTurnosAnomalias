"""
ETL de normalizacion de la malla de turnos (export SERPI) hacia PostgreSQL.

Uso:
    python etl_normalizacion.py --archivo RepProgramacion.xlsx

Requiere un archivo .env (ver .env.example) con los datos de conexion a
PostgreSQL. Es idempotente: se puede correr varias veces con el mismo
archivo sin duplicar turnos (usa upsert sobre una llave natural).
"""

import argparse
import os
import re
import sys
from datetime import datetime

import openpyxl
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv

DAY_RE = re.compile(r'^(\d{1,2})\n')
TIME_RE = re.compile(r'^\d{1,2}:\d{2}$')
NAME_RE = re.compile(r'^(.*?)\s*-\s*(\d{5,12})\s*(?:-\s*(.*))?$')
HOURS_RE = re.compile(r'(Horas [A-Za-zÁÉÍÓÚñ ]+|Total Horas):\s*(\d+)')

MESES = {
    'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4, 'Mayo': 5, 'Junio': 6,
    'Julio': 7, 'Agosto': 8, 'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11,
    'Diciembre': 12,
}

LEAVE_CODES = {'VAC', 'INC', 'LIC', 'LICNR'}


# --------------------------------------------------------------------------
# Parsing del Excel (misma logica validada sobre la malla de julio 2026)
# --------------------------------------------------------------------------

def build_day_map(ws, header_row, max_col):
    dmap = {}
    for c in range(1, max_col + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str):
            m = DAY_RE.match(v)
            if m:
                dmap[int(m.group(1))] = c
    return dmap


def parse_shift_cell(v):
    """Devuelve (codigo, inicio, fin) o None si la celda esta vacia."""
    if not isinstance(v, str) or not v.strip():
        return None
    parts = [p.strip() for p in v.split('\n') if p.strip() != '']
    if len(parts) == 1:
        return (parts[0], None, None)
    elif len(parts) == 2:
        if TIME_RE.match(parts[0]) and TIME_RE.match(parts[1]):
            return (None, parts[0], parts[1])  # celda sin letra de codigo
        return (parts[0], parts[1], None)
    else:
        return (parts[0], parts[1], parts[2])


def duration_hours(start, end):
    if not start or not end:
        return None
    sh, sm = map(int, start.split(':'))
    eh, em = map(int, end.split(':'))
    s, e = sh * 60 + sm, eh * 60 + em
    if e <= s:
        e += 24 * 60
    return round((e - s) / 60.0, 2)


def parse_name(name_raw):
    m = NAME_RE.match(name_raw)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return name_raw.strip(), None


def find_year_month(ws, max_col):
    for r in range(1, 15):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        for c, v in enumerate(row_vals):
            if isinstance(v, str) and v.strip() == 'Año:':
                anio = row_vals[c + 5] if c + 5 < len(row_vals) else None
                # el mes esta unas celdas mas a la derecha en la misma fila
                for cc in range(c, len(row_vals)):
                    vv = row_vals[cc]
                    if isinstance(vv, str) and vv.strip() == 'Mes:':
                        for ccc in range(cc, len(row_vals)):
                            mes_val = row_vals[ccc]
                            if isinstance(mes_val, str) and mes_val.strip() in MESES:
                                return int(anio), MESES[mes_val.strip()]
    return None, None


def parse_workbook(path):
    """Devuelve una lista de dicts, uno por asignacion guarda-puesto, con sus
    turnos dia a dia y las horas declaradas en el encabezado."""
    wb = openpyxl.load_workbook(path, data_only=True)
    records = []

    for sname in wb.sheetnames:
        ws = wb[sname]
        max_row, max_col = ws.max_row, ws.max_column

        anio, mes = find_year_month(ws, max_col)

        cliente = proyecto = puesto = None
        r = 1
        day_map = {}
        while r <= max_row:
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
            marker_text = None
            for v in row_vals:
                if isinstance(v, str) and v.strip().startswith(('CLIENTE:', 'PROYECTO:', 'PUESTO:')):
                    marker_text = v.strip()
                    break
            if marker_text is not None:
                if marker_text.startswith('CLIENTE:'):
                    cliente = marker_text[len('CLIENTE:'):].strip()
                elif marker_text.startswith('PROYECTO:'):
                    proyecto = marker_text[len('PROYECTO:'):].strip()
                elif marker_text.startswith('PUESTO:'):
                    puesto = marker_text[len('PUESTO:'):].strip()
                    header_row = None
                    for rr in range(r + 1, min(r + 4, max_row + 1)):
                        if ws.cell(row=rr, column=3).value == 'T':
                            header_row = rr
                            break
                    if header_row:
                        day_map = build_day_map(ws, header_row, max_col)
                        r = header_row + 1
                        continue
                r += 1
                continue

            c3 = ws.cell(row=r, column=3).value
            c4 = ws.cell(row=r, column=4).value
            if isinstance(c3, int) and isinstance(c4, str) and c4.strip() and day_map:
                name, cedula = parse_name(c4.strip())
                hours_decl = {}
                for c in range(1, max_col + 1):
                    v = ws.cell(row=r, column=c).value
                    if isinstance(v, str):
                        m = HOURS_RE.match(v.strip())
                        if m:
                            hours_decl[m.group(1)] = int(m.group(2))

                shift_row = r + 1
                turnos = []
                if shift_row <= max_row and cedula:
                    for day, col in day_map.items():
                        parsed = parse_shift_cell(ws.cell(row=shift_row, column=col).value)
                        if not parsed:
                            continue
                        code, start, end = parsed
                        dur = duration_hours(start, end)
                        if code is None:
                            # celda sin letra: codigo sintetico a partir del horario
                            code = f"T_{start.replace(':','')}_{end.replace(':','')}"
                        categoria = 'AUSENCIA' if code in LEAVE_CODES else 'TRABAJADO'
                        turnos.append({
                            'dia': day, 'codigo': code, 'inicio': start, 'fin': end,
                            'horas': dur, 'categoria': categoria,
                        })

                if cedula:
                    records.append({
                        'sheet': sname, 'cliente': cliente, 'puesto': puesto,
                        'guarda_cedula': cedula, 'guarda_nombre': name, 'slot': c3,
                        'anio': anio, 'mes': mes,
                        'hours_decl': hours_decl, 'turnos': turnos,
                    })
                r = shift_row + 1
                continue
            r += 1
    return records


# --------------------------------------------------------------------------
# Insercion en PostgreSQL
# --------------------------------------------------------------------------

def get_connection():
    load_dotenv()
    return psycopg2.connect(
        host=os.environ['DB_HOST'],
        port=os.environ.get('DB_PORT', '5432'),
        dbname=os.environ['DB_NAME'],
        user=os.environ['DB_USER'],
        password=os.environ['DB_PASSWORD'],
        sslmode=os.environ.get('DB_SSLMODE', 'prefer'),
    )


def load(records, conn):
    cur = conn.cursor()

    cliente_id = {}
    puesto_id = {}
    guardas_vistos = set()
    tipos_vistos = {}

    n_turnos = 0
    n_horas_decl = 0

    for rec in records:
        cliente = rec['cliente'] or '(sin cliente)'
        if cliente not in cliente_id:
            cur.execute(
                """INSERT INTO clientes (nombre) VALUES (%s)
                   ON CONFLICT (nombre) DO UPDATE SET nombre = EXCLUDED.nombre
                   RETURNING id""",
                (cliente,),
            )
            cliente_id[cliente] = cur.fetchone()[0]

        puesto = rec['puesto'] or '(sin puesto)'
        pkey = (cliente, puesto)
        if pkey not in puesto_id:
            cur.execute(
                """INSERT INTO puestos (cliente_id, nombre, hoja_origen) VALUES (%s, %s, %s)
                   ON CONFLICT (cliente_id, nombre)
                   DO UPDATE SET hoja_origen = EXCLUDED.hoja_origen
                   RETURNING id""",
                (cliente_id[cliente], puesto, rec['sheet']),
            )
            puesto_id[pkey] = cur.fetchone()[0]

        cedula = rec['guarda_cedula']
        if cedula not in guardas_vistos:
            cur.execute(
                """INSERT INTO guardas (cedula, nombre) VALUES (%s, %s)
                   ON CONFLICT (cedula) DO UPDATE SET nombre = EXCLUDED.nombre""",
                (cedula, rec['guarda_nombre']),
            )
            guardas_vistos.add(cedula)

        this_puesto_id = puesto_id[pkey]

        for t in rec['turnos']:
            tkey = t['codigo']
            if tkey not in tipos_vistos:
                cur.execute(
                    """INSERT INTO tipos_turno (codigo, hora_inicio, hora_fin, duracion_horas, categoria)
                       VALUES (%s, %s, %s, %s, %s)
                       ON CONFLICT (codigo) DO NOTHING""",
                    (t['codigo'], t['inicio'], t['fin'], t['horas'], t['categoria']),
                )
                tipos_vistos[tkey] = True

            if not (rec['anio'] and rec['mes']):
                continue
            fecha = f"{rec['anio']:04d}-{rec['mes']:02d}-{t['dia']:02d}"

            cur.execute(
                """INSERT INTO turnos
                       (guarda_cedula, puesto_id, fecha, slot, tipo_turno_codigo,
                        hora_inicio, hora_fin, horas_calculadas, origen)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'EXCEL_EXPORT')
                   ON CONFLICT (guarda_cedula, puesto_id, fecha, slot)
                   DO UPDATE SET tipo_turno_codigo = EXCLUDED.tipo_turno_codigo,
                                 hora_inicio = EXCLUDED.hora_inicio,
                                 hora_fin = EXCLUDED.hora_fin,
                                 horas_calculadas = EXCLUDED.horas_calculadas,
                                 fecha_carga = now()""",
                (cedula, this_puesto_id, fecha, rec['slot'], t['codigo'],
                 t['inicio'], t['fin'], t['horas']),
            )
            n_turnos += 1

        hd = rec['hours_decl']
        if hd and rec['anio'] and rec['mes']:
            cur.execute(
                """INSERT INTO horas_declaradas_mes
                       (guarda_cedula, puesto_id, slot, anio, mes,
                        horas_diurnas_ordinarias, horas_diurnas_festivas,
                        horas_nocturnas_ordinarias, horas_nocturnas_festivas, total_horas)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                   ON CONFLICT (guarda_cedula, puesto_id, slot, anio, mes)
                   DO UPDATE SET
                       horas_diurnas_ordinarias = EXCLUDED.horas_diurnas_ordinarias,
                       horas_diurnas_festivas = EXCLUDED.horas_diurnas_festivas,
                       horas_nocturnas_ordinarias = EXCLUDED.horas_nocturnas_ordinarias,
                       horas_nocturnas_festivas = EXCLUDED.horas_nocturnas_festivas,
                       total_horas = EXCLUDED.total_horas""",
                (cedula, this_puesto_id, rec['slot'], rec['anio'], rec['mes'],
                 hd.get('Horas Diurnas Ordinarias'), hd.get('Horas Diurnas Festivas'),
                 hd.get('Horas Nocturnas Ordinarias'), hd.get('Horas Nocturnas Festivas'),
                 hd.get('Total Horas')),
            )
            n_horas_decl += 1

    conn.commit()
    cur.close()
    return {
        'clientes': len(cliente_id), 'puestos': len(puesto_id),
        'guardas': len(guardas_vistos), 'tipos_turno': len(tipos_vistos),
        'turnos': n_turnos, 'horas_declaradas': n_horas_decl,
    }


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument('--archivo', required=True, help='Ruta al Excel exportado de SERPI')
    args = ap.parse_args()

    print(f"Leyendo {args.archivo} ...")
    records = parse_workbook(args.archivo)
    print(f"  {len(records)} asignaciones guarda-puesto encontradas")

    sin_fecha = sum(1 for r in records if not (r['anio'] and r['mes']))
    if sin_fecha:
        print(f"  ADVERTENCIA: {sin_fecha} hojas sin Año/Mes detectado, sus turnos no se insertaron")

    print("Conectando a PostgreSQL ...")
    conn = get_connection()

    print("Insertando (upsert, se puede correr varias veces sin duplicar) ...")
    stats = load(records, conn)
    conn.close()

    print("Listo:")
    for k, v in stats.items():
        print(f"  {k}: {v}")


if __name__ == '__main__':
    sys.exit(main())
